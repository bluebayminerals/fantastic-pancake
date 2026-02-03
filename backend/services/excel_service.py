from openpyxl import Workbook, load_workbook
from io import BytesIO
from typing import List, Dict
from datetime import datetime

class ExcelService:
    @staticmethod
    def generate_shops_template() -> BytesIO:
        """Generate Excel template for shop import"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Shops"
        
        # Headers
        headers = ["Name", "Address", "Phone", "Contact Person", "Pincode", 
                  "Latitude", "Longitude", "Delivery Time (mins)"]
        ws.append(headers)
        
        # Sample data
        ws.append(["Sample Shop", "123 Main St, Kochi", "9876543210", "John Doe", 
                  "682001", "9.9312", "76.2673", "15"])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def parse_shops_excel(file_content: bytes) -> List[Dict]:
        """Parse uploaded Excel file and return shop data"""
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active
        
        shops = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Check if name exists
                shop = {
                    "name": row[0],
                    "address": row[1],
                    "phone": str(row[2]),
                    "contact_person": row[3],
                    "pincode": str(row[4]),
                    "latitude": float(row[5]) if row[5] else 0,
                    "longitude": float(row[6]) if row[6] else 0,
                    "delivery_time_minutes": int(row[7]) if row[7] else 15
                }
                shops.append(shop)
        
        return shops
    
    @staticmethod
    def generate_sales_report(sales_data: List[Dict]) -> BytesIO:
        """Generate Excel report for sales"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Headers
        headers = ["Date", "Shop", "Salesman", "Total Amount", "Payment Mode", "Items Count"]
        ws.append(headers)
        
        # Data
        for sale in sales_data:
            ws.append([
                sale.get("sale_date", ""),
                sale.get("shop_name", ""),
                sale.get("salesman_name", ""),
                sale.get("total_amount", 0),
                sale.get("payment_mode", ""),
                len(sale.get("items", []))
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

excel_service = ExcelService()